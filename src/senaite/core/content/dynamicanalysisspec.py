# -*- coding: utf-8 -*-
#
# This file is part of SENAITE.CORE.
#
# SENAITE.CORE is free software: you can redistribute it and/or modify it under
# the terms of the GNU General Public License as published by the Free Software
# Foundation, version 2.
#
# This program is distributed in the hope that it will be useful, but WITHOUT
# ANY WARRANTY; without even the implied warranty of MERCHANTABILITY or FITNESS
# FOR A PARTICULAR PURPOSE. See the GNU General Public License for more
# details.
#
# You should have received a copy of the GNU General Public License along with
# this program; if not, write to the Free Software Foundation, Inc., 51
# Franklin Street, Fifth Floor, Boston, MA 02110-1301 USA.
#
# Copyright 2018-2024 by it's authors.
# Some rights reserved, see README and LICENSE.

from collections import defaultdict

from bika.lims import senaiteMessageFactory as _
from bika.lims import api
from senaite.core.catalog import SETUP_CATALOG
from senaite.core.content.base import Container
from senaite.core.interfaces import IDynamicAnalysisSpec
from openpyxl.reader.excel import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from plone.namedfile import field as namedfile
from plone.supermodel import model
from six import StringIO
from z3c.form.interfaces import NOT_CHANGED
from zope.interface import Invalid
from zope.interface import implementer
from zope.interface import invariant

REQUIRED_COLUMNS = [
    "Keyword",  # The Analysis Keyword
    "min",  # Lower Limit
    "max",  # Upper Limit
]


class IDynamicAnalysisSpecSchema(model.Schema):
    """Dynamic Analysis Specification
    """

    specs_file = namedfile.NamedBlobFile(
        title=_(u"Specification File"),
        description=_(u"Only Excel files supported"),
        required=True)

    @invariant
    def validate_specs_file(data):
        """Checks the Excel file contains the required header columns
        """
        # return immediately if not changed
        if data.specs_file == NOT_CHANGED:
            return True
        fd = StringIO(data.specs_file.data)
        try:
            xls = load_workbook(fd)
        except (InvalidFileException, TypeError):
            raise Invalid(_(
                "Invalid specifications file detected. "
                "Please upload an Excel spreadsheet with at least "
                "the following columns defined: '{}'"
                .format(", ".join(REQUIRED_COLUMNS))))
        try:
            header_row = xls.worksheets[0].rows.next()
            header = map(lambda c: c.value, header_row)
        except (IndexError, AttributeError):
            raise Invalid(
                _("First sheet does not contain a valid column definition"))
        for col in REQUIRED_COLUMNS:
            if col not in header:
                raise Invalid(_("Column '{}' is missing".format(col)))


@implementer(IDynamicAnalysisSpec, IDynamicAnalysisSpecSchema)
class DynamicAnalysisSpec(Container):
    """Dynamic Analysis Specification
    """
    _catalogs = [SETUP_CATALOG]

    def get_workbook(self):
        specs_file = self.specs_file
        if not specs_file:
            return None
        data = StringIO(specs_file.data)
        return load_workbook(data)

    def get_worksheets(self):
        wb = self.get_workbook()
        if wb is None:
            return []
        return wb.worksheets

    def get_primary_sheet(self):
        sheets = self.get_worksheets()
        if len(sheets) == 0:
            return None
        return sheets[0]

    def get_header(self):
        header = []
        ps = self.get_primary_sheet()
        if ps is None:
            return header
        for num, row in enumerate(ps.rows):
            if num > 0:
                break
            header = [cell.value for cell in row]
        return header

    def get_specs(self):
        ps = self.get_primary_sheet()
        if ps is None:
            return []
        keys = self.get_header()
        specs = []

        def get_cell_string_value(cell):
            value = cell.value
            if api.is_string(value):
                return value
            elif value is None:
                return None
            return str(value)

        for num, row in enumerate(ps.rows):
            # skip the header
            if num == 0:
                continue
            values = map(get_cell_string_value, row)
            data = dict(zip(keys, values))
            specs.append(data)
        return specs

    def get_by_keyword(self):
        specs = self.get_specs()
        groups = defaultdict(list)
        for spec in specs:
            groups[spec.get("Keyword")].append(spec)
        return groups
